from conversion import Conversion
from transaction import Buy, Receive, Sell, Send, Receive, Transaction
import pandas as pd
import os
from ast import literal_eval
from time import strftime
from openpyxl import load_workbook, Workbook, workbook
import base64
import datetime
import io
from openpyxl import load_workbook, Workbook
import pandas as pd
from time import strftime
import os
from ast import literal_eval

import numpy as np
from assets import Asset
import sys
import pathlib

from utils import *

# basedir = os.path.abspath(os.path.dirname(__file__))
basedir = os.path.abspath(os.path.dirname(sys.executable))

class Transactions:

    def __init__(self):
        
        self.saves = self.load_saves()
        self.index = 0
        self.conversions = []
        self.asset_objects = []

        if len(self.saves) > 0:
            view = self.saves[-1]['value']
            self.transactions = self.load(view)
            self.view = view
        else:
            self.view = ""
            self.transactions = []

    def __len__(self):
        return len(self.transactions)

    def __iter__(self):
        self.index = 0
        return self
    
    def __next__(self):
        try:
            result = self.transactions[self.index]
        except IndexError:
            raise StopIteration
        
        self.index += 1
        return result

    @property
    def links(self):
        links = set([
                link 
                for trans in self.transactions
                for link in trans.links
                ])
        
        return links

    def link_with_fifo(self, asset=None, min_link=0.01):
        
        sells = {}
        buys = {}

        # Sort into Buys a Sells
        for trans in self.transactions:

            # If asset is provided only auto-link symbol provided
            if asset is not None:
                if trans.symbol != asset:
                    continue

            if trans.trans_type == 'buy':

                if trans.symbol not in buys.keys():
                    buys[trans.symbol] = []

                buys[trans.symbol].append(trans)

            elif trans.trans_type == 'sell':
                
                if trans.symbol not in sells.keys():
                    sells[trans.symbol] = []

                sells[trans.symbol].append(trans)
        
        # Sort By Time Stamp
        for key in buys.keys():
            
            buys[key].sort(key=lambda x: x.time_stamp)

        for key in sells.keys():
            
            sells[key].sort(key=lambda x: x.time_stamp)


        # Start the Algo
        for key in sells.keys():

            for sell in sells[key]:
 
                # check if sell has remaining unlinked quantity
                if sell.unlinked_quantity > 0:
                    
                    for buy in buys[key]:

                        # Skip if buy has no remaining unlinked quantity
                        if buy.unlinked_quantity <= 0:
                            continue
                        
                        # break if sell has no remaining unlinked quantity
                        if sell.unlinked_quantity <= 0:
                            break
                        
                        # check if buy came before sell
                        if buy.time_stamp >= sell.time_stamp:
                            continue

                        # Link 
                        # if sell unlinked is greater than buy unlinked, link quantity equals buy unlinked
                        if sell.unlinked_quantity >= buy.unlinked_quantity:
                            link_quantity = buy.unlinked_quantity
                        
                        # if sell unlinked is less than buy unlinked, link quantity equals sell unlinked
                        elif sell.unlinked_quantity <= buy.unlinked_quantity: 
                            link_quantity = sell.unlinked_quantity
                        
                        
                        # if sell.unlinked_quantity <= min_link:
                        #     print(f"    [{sell.unlinked_quantity}] <=  [{min_link}] MIN LINK SURPASSED BEFORE LINK  {link_quantity}  !!!!!")

                        # Determine link profitability
                        buy_price = link_quantity * buy.usd_spot
                        sell_price = link_quantity * sell.usd_spot
                        profit = sell_price - buy_price

                        if abs(profit) < 0.01:
                            continue
                        
                        sell.link_transaction(buy, link_quantity)
                        
                    
                        # if sell.unlinked_quantity <= min_link:
                        #     print(f"    [{sell.unlinked_quantity}] <=  [{min_link}] MIN LINK SURPASSED AFTER LINK  {link_quantity}  !!!!!")
                            
                        
                        

    def link_with_filo(self, asset):

        sells = {}
        buys = {}

        # Sort into Buys a Sells
        for trans in self.transactions:

            # If asset is provided only auto-link symbol provided
            if asset is not None:
                if trans.symbol != asset:
                    continue

            if trans.trans_type == 'buy':

                if trans.symbol not in buys.keys():
                    buys[trans.symbol] = []

                buys[trans.symbol].append(trans)

            elif trans.trans_type == 'sell':
                
                if trans.symbol not in sells.keys():
                    sells[trans.symbol] = []

                sells[trans.symbol].append(trans)
        
        # Sort By Time Stamp
        for key in buys.keys():

            buys[key].sort(key=lambda x: x.time_stamp, reverse=True)
        
        for key in sells.keys():
            
            sells[key].sort(key=lambda x: x.time_stamp)

        # Start the Algo
        for key in sells.keys():

            for sell in sells[key]:

                # check if sell has remaining unlinked quantity
                if sell.unlinked_quantity > 0:
                    
                    for buy in buys[key]:
                        
                        # Skip if buy has no remaining unlinked quantity
                        if buy.unlinked_quantity == 0:
                            continue
                        
                        # break if sell has no remaining unlinked quantity
                        if sell.unlinked_quantity == 0:
                            break
                        
                        # check if sell came after buy
                        if sell.time_stamp <= buy.time_stamp:
                            continue

                                                
                        # if sell unlinked is greater than buy unlinked, link quantity equals buy unlinked
                        if sell.unlinked_quantity >= buy.unlinked_quantity:
                            link_quantity = buy.unlinked_quantity
                        
                        # if sell unlinked is less than buy unlinked, link quantity equals sell unlinked
                        elif sell.unlinked_quantity <= buy.unlinked_quantity: 
                            link_quantity = sell.unlinked_quantity
                        
                        # Determine link profitability
                        buy_price = link_quantity * buy.usd_spot
                        sell_price = link_quantity * sell.usd_spot
                        profit = sell_price - buy_price

                        if abs(profit) < 0.01:
                            continue
                        
                        # Link 
                        sell.link_transaction(buy, link_quantity)

        

    
    def convert_buys_to_lost(self, asset, amount):
        # method used to deal with crypto not sold on exchange but no longer in possession

        if asset in self.assets:
            
            # list of buys that were converted to lost
            buys_to_delete = []
            
            # what we want to calc
            bought = 0.0
            quantity_of_buys_converted_to_lost = 0.0

            for trans in self.transactions:
                if trans.symbol != asset:
                    continue

                if trans.trans_type == 'buy':
                    bought += trans.quantity



            
            for trans in self.transactions:
                if trans.symbol != asset:
                    continue
                
                if amount > 0 and trans.quantity <= amount and trans.trans_type == 'buy':
                    # Convert buy to lost
                  
                    conversion = Conversion(input_trans_type='buy', 
                                            output_trans_type='lost', 
                                            input_symbol=asset, 
                                            input_quantity=trans.quantity, 
                                            input_time_stamp=trans.time_stamp, 
                                            input_usd_spot=trans.usd_spot, 
                                            input_usd_total=trans.usd_total, 
                                            reason="Current HODL Buy to lost/deleted")
                    
                    self.conversions.append(conversion)
                
                    buys_to_delete.append(trans)

                    quantity_of_buys_converted_to_lost += trans.quantity


        for trans in self.transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()

        
        print(f" Num of transactions to delete {len(buys_to_delete)} ")
        print(f" Num of transactions before delete {len(self.transactions)} ")
        

        for trans in buys_to_delete:
            self.transactions.remove(trans)

    
        print(f" Num of transactions after delete {len(self.transactions)} ")
        
        # what we want to calc a second time
        bought_after = 0.0

        for trans in self.transactions:
            if trans.symbol != asset:
                continue

            if trans.trans_type == 'buy':
                bought_after += trans.quantity

    
        print(f"AFTER CONVERSION bought {bought_after} {asset}")


    

    def convert_sends_to_sells(self, asset, amount_to_convert):
        # method used to deal with crypto not sold on exchange but no longer in possession
        ##### Sort these

        if asset in self.assets:
            
            # list of sends that were converted to sells
            sends = []
            sells = []
            sends_to_delete = []

            # what we want to calc
            bought = 0.0
            sold = 0.0
            sent = 0.0
            quantity_of_sends_converted_to_sells = 0.0
            inital_amount_to_convert = amount_to_convert

            self.transactions.sort(key=lambda x: x.time_stamp)

            for trans in self.transactions:
                if trans.symbol != asset:
                    continue

                if trans.trans_type != 'send':
                    continue

                if amount_to_convert > 0 and trans.quantity <= amount_to_convert:
                    # Convert Send to Sell
                    # print(f"converting Send {trans.time_stamp} {trans.quantity} to sell")

                    if type(trans.usd_spot) != float:
                        print(f"skipping transaction {trans.name} because no usd_spot")
                        continue
                    
                    sell = Sell(symbol=trans.symbol, quantity=trans.quantity, time_stamp=trans.time_stamp, usd_spot=trans.usd_spot, linked_transactions=trans.linked_transactions, source="Gainz App Send to Sale")
                
                    conversion = Conversion(input_trans_type='send', 
                                            output_trans_type='sell', 
                                            input_symbol=asset, 
                                            input_quantity=trans.quantity, 
                                            input_time_stamp=trans.time_stamp, 
                                            input_usd_spot=trans.usd_spot, 
                                            input_usd_total=trans.usd_total, 
                                            reason="Current HODL Send to Sell", 
                                            source=f"{trans.source} Converted in Gainz App")

                    self.conversions.append(conversion)
                    
                    self.transactions.append(sell)

                    sends_to_delete.append(trans)

                    quantity_of_sends_converted_to_sells += trans.quantity
                    
                    amount_to_convert -= trans.quantity

        for trans in self.transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()
        
        print(f"\n\nSuccessfully Converted {quantity_of_sends_converted_to_sells} in Sends to Sells in {len(sends_to_delete)} transactions!!\n\n")
        print(f" Num of transactions to delete {len(sends_to_delete)} ")
        print(f" Num of transactions before delete {len(self.transactions)} ")
        

        for trans in sends_to_delete:
            self.transactions.remove(trans)

        # del sends 
        # del sells

        print(f" Num of transactions after delete {len(self.transactions)} ")
        
    
        # what we want to calc a second time
        bought_after = 0.0
        sold_after = 0.0
        sent_after = 0.0

        for trans in self.transactions:
            if trans.symbol != asset:
                continue

            if trans.trans_type == 'buy':
                bought_after += trans.quantity

            elif trans.trans_type == 'sell':
                sold_after += trans.quantity
            
            elif trans.trans_type == 'send':
                sent_after += trans.quantity

        print(f"AFTER CONVERSION Sold {sold_after} {asset}")
        print(f"AFTER CONVERSION Sent {sent_after} {asset}")
        print(f"AFTER CONVERSION Unaccounted for {amount_to_convert} {asset}")
    
     

    def load_saves(self):

        saves = []

        match_object = "saved_"
        view_num = 1
        for root, dirs, files in os.walk('./saves'):
            for f in files:
                if match_object in f and f.endswith('xlsx'):
                    workbook = load_workbook(filename=f"./saves/{f}")
                    if 'Description' in workbook.sheetnames:
                        sheet = workbook['Description']
                        description = sheet.cell(column=1, row=1).value
                    else:
                        description = ""
                    saves.append({'label': f"./saves/{f}", 'value': f"./saves/{f}", 'description': description})
                    view_num += 1

        self.saves = saves
        
        return saves

    @property    
    def assets(self):

        assets = set()
    
        for trans in self.transactions:
            assets.add(trans.symbol)

        return assets

    # Load Previous Data returns view options
    def load(self, filename=None):

        # Read Previously saved data into pandas df - Transactions
        trans_df = pd.read_excel(filename, sheet_name='All Transactions', converters = {'my_str_column': list})
        trans_df['linked_transactions'] = trans_df['linked_transactions'].apply(lambda x: literal_eval(str(x)))
        trans_df.reset_index(inplace=True)
        

        # Read Previously saved data into pandas df - Conversions
        conversion_df = pd.read_excel(filename, sheet_name='Conversions', converters = {'my_str_column': list})
        conversion_df.reset_index(inplace=True)

        # Read Previously saved data into pandas df - Assets
        asset_df = pd.read_excel(filename, sheet_name='Assets', converters = {'my_str_column': list})


        # Read Previously saved data into pandas df - Links
        links_df = pd.read_excel(filename, sheet_name='Links', converters = {'my_str_column': list})

        # Split Buys and Sells into separate df's
        sell_df = trans_df[(trans_df['trans_type'] == 'sell')].copy()
        buy_df = trans_df[(trans_df['trans_type'] == 'buy')].copy()
        send_df = trans_df[(trans_df['trans_type'] == 'send')].copy()
        receive_df = trans_df[(trans_df['trans_type'] == 'receive')].copy()
        
        send_df.reset_index(inplace=True)
        sell_df.reset_index(inplace=True)
        buy_df.reset_index(inplace=True)
        receive_df.reset_index(inplace=True)
        
        sell_df.sort_values(by='time_stamp', inplace=True)
        buy_df.sort_values(by='time_stamp', inplace=True)
        send_df.sort_values(by='time_stamp', inplace=True)
        receive_df.sort_values(by='time_stamp', inplace=True)

        # Objects >
        sells = []
        buys = []
        sends = []
        receives = []
        conversions = []
        asset_objects = []


        # Load Transactions into Objects
        
        # Load Sells
        for index, row in sell_df.iterrows():
            sells.append(Sell(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], linked_transactions=row['linked_transactions'], source=row['source']))
        
        # Load Buys
        for index, row in buy_df.iterrows():
            buys.append(Buy(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], linked_transactions=row['linked_transactions'], source=row['source']))
        
        # Load Sends
        for index, row in send_df.iterrows():
            # buys.append(Buy(symbol='BTC', quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction']))
            sends.append(Send(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], linked_transactions=row['linked_transactions'], source=row['source']))

        # Load Receives
        for index, row in receive_df.iterrows():
            # buys.append(Buy(symbol='BTC', quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction']))
            receives.append(Receive(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], linked_transactions=row['linked_transactions'], source=row['source']))
        
        # Load Conversions
        for index, row in conversion_df.iterrows():
            conversions.append(Conversion(
                input_trans_type=row['input_trans_type'],
                output_trans_type=row['output_trans_type'],
                input_symbol=row['symbol'],
                input_quantity=row['quantity'],
                input_time_stamp=row['time_stamp'],
                input_usd_spot=row['usd_spot'],
                input_usd_total=row['usd_total'],
                reason=row['reason'],
                source=row['source']               
                )
            )

        # load Assets
        for index, row in asset_df.iterrows():
            asset_objects.append(Asset(symbol=row['symbol'], hodl=row['hodl']))

        self.asset_objects = asset_objects

        imported_transactions = buys + sells + sends + receives

        # Duplicates check
        transactions = set()
        for trans in imported_transactions:
            transactions.add(trans)

        # Multi-Link Indicator
        for trans in transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()

        self.transactions = list(transactions)
        self.conversions = conversions

        # Re-import Re-Create Links
        for index, row in links_df.iterrows():
            buy = row['buy'].strip("'")
            sell = row['sell'].strip("'")
            symbol = row['symbol']
            quantity = row['quantity']

            buy_obj = None
            sell_obj = None
            
            for trans in self.transactions:

                if trans.name == sell:
                    # print(f'Sell Found {trans.name}')
                    sell_obj = trans

                elif trans.name == buy:
                    # print(f'Buy Found {trans.name}')
                    buy_obj = trans

                if buy_obj and sell_obj:
                    # print("Buy and Sell Found Breaking\n")
                    break
            
            if sell_obj and buy_obj:
                # print(f" quantity of link on load {quantity}")
                sell_obj.link_transaction(buy_obj, link_quantity=quantity)


        return list(transactions)


    def filter(self, symbols, options):

        trans_in_view = set()
        
        for trans in self.transactions:

            # if symbol in view
            if trans.symbol in symbols:
            
                # if has_link in options
                if 'has_link' in options:
                    if len(trans.links) > 0:
                        trans_in_view.add(trans)
                
                if 'needs_link' in options and trans.trans_type == 'sell':
                    if trans.unlinked_quantity > 0.0:
                        trans_in_view.add(trans)

                if 'no_link' in options:
                    if len(trans.links) == 0:
                        trans_in_view.add(trans)
                        

        return list(trans_in_view)


    def save(self, description=None):
        save_as_filename = f"./saves/saved_{strftime('Y%Y-M%m-D%d_H%H-M%M-S%S')}.xlsx"
        
        for trans in self.transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()
        
        trans_df = pd.DataFrame([vars(s) for s in self.transactions])
        conversion_df = pd.DataFrame([vars(s) for s in self.conversions])
        asset_df = pd.DataFrame([vars(s) for s in self.asset_objects])

        # for link in self.links:
        #     print(f"Link Quantity before save {link.quantity}")


        links_df = pd.DataFrame([vars(s) for s in self.links])

        # for link in self.links:
            # print(f"\nQuantity of link on Save {link.quantity} \n   buy quantity {link.buy.quantity} unlinked {link.buy.unlinked_quantity} \n   sell quantity {link.sell.quantity} unlinked {link.sell.unlinked_quantity}")


        with pd.ExcelWriter(save_as_filename) as writer:
            trans_df.to_excel(writer, sheet_name="All Transactions")
            conversion_df.to_excel(writer, sheet_name="Conversions")
            asset_df.to_excel(writer, sheet_name="Assets")
            links_df.to_excel(writer, sheet_name="Links")

        workbook = load_workbook(filename=save_as_filename)
        sheet = workbook.create_sheet('Description')

        sheet.cell(row=1, column=1, value=description)

        workbook.save(save_as_filename)
        workbook.close()

        print(f"{description} Saving to {save_as_filename}")

        self.saves = self.load_saves()
        self.view = self.saves[-1]['value']
        
        return save_as_filename


    def delete(self, filename):
        os.rename(filename, f"{filename}.bak")


    def export_to_excel(self, asset=None, date_range=None, by_year=True):

        # Idea to programatically create Excel Links, Fancy ;-)
        # =HYPERLINK("[Export_Y2021-M03-D06_H19-M34.xlsx]Links!A20","Display Text")
        
        save_as_filename = f"./Exports/Export_{strftime('Y%Y-M%m-D%d_H%H-M%M')}.xlsx"
        
        # Template to use
        workbook = load_workbook(filename='Gainz_Template.xlsx')
        c_sheet = workbook['Conversions']
        l_sheet = workbook['Gains']
        a_sheet = workbook['All Transactions']
        s_sheet = workbook['Stats']
        t8949_sheet = workbook['8949']



        # 8949 Short
        years = set()
        for link in self.links:

            years.add(link.sell.time_stamp.year)

        for year in years:
            
            sheetname = f'{year} 8949 Short'
            ws = workbook.copy_worksheet(t8949_sheet)
            ws.title = sheetname
            
            row = 2
            for link in self.links:

                if link.sell.time_stamp.year != year:
                    continue

                if abs(link.profit_loss) <= 1:
                    continue
                
                if link.hodl_duration.days > 365:
                    continue

                ws[f"A{row}"] = f"Crypto {link.symbol}"
                ws[f"B{row}"] = link.buy.time_stamp
                ws[f"C{row}"] = link.sell.time_stamp
                ws[f"D{row}"] = link.link_sell_price
                ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                ws[f"E{row}"] = link.link_buy_price
                ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                ws[f"H{row}"] = link.profit_loss
                ws[f"H{row}"].number_format = '"$"#,##0.00_-'

                row += 1

            if row == 2:
                workbook.remove(ws)

            else:
                row += 2

                ws[f"C{row}"] = "Totals"

                ws[f"D{row}"] = f"=SUM(D2:D{row -2})"
                ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                ws[f"E{row}"] = f"=SUM(E2:E{row -2})"
                ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                ws[f"H{row}"] = f"=SUM(H2:H{row -2})"
                ws[f"H{row}"].number_format = '"$"#,##0.00_-'


        # 8949 Long
        years = set()
        for link in self.links:

            years.add(link.sell.time_stamp.year)

        for year in years:
            
            sheetname = f'{year} 8949 Long'
            ws = workbook.copy_worksheet(t8949_sheet)
            ws.title = sheetname
            
            row = 2
            for link in self.links:

                if link.sell.time_stamp.year != year:
                    continue

                if abs(link.profit_loss) <= 1:
                    continue
                
                if link.hodl_duration.days <= 365:
                    continue

                ws[f"A{row}"] = f"Crypto {link.symbol}"
                ws[f"B{row}"] = link.buy.time_stamp
                ws[f"C{row}"] = link.sell.time_stamp
                ws[f"D{row}"] = link.link_sell_price
                ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                ws[f"E{row}"] = link.link_buy_price
                ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                ws[f"H{row}"] = link.profit_loss
                ws[f"H{row}"].number_format = '"$"#,##0.00_-'

                row += 1

            if row == 2:
                workbook.remove(ws)

            else:
                row += 2

                ws[f"C{row}"] = "Totals"

                ws[f"D{row}"] = f"=SUM(D2:D{row -2})"
                ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                ws[f"E{row}"] = f"=SUM(E2:E{row -2})"
                ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                ws[f"H{row}"] = f"=SUM(H2:H{row -2})"
                ws[f"H{row}"].number_format = '"$"#,##0.00_-'


        for asset in self.assets:
            
            # Conversions sheet
            sheetname = f'{asset} Conversions'
            conversions_sheet = workbook.copy_worksheet(c_sheet)
            conversions_sheet.title = sheetname

            column_names = []
            for cell in conversions_sheet[3]:
                column_names.append(cell.value)
            
            in_trans_type_index = column_names.index("In Transaction Type") + 1
            out_trans_type_index = column_names.index("Out Transaction Type") + 1
            symbol_index = column_names.index("Symbol") + 1
            time_stamp_index = column_names.index("Time Stamp") + 1
            quantity_index = column_names.index("Quantity") + 1
            usd_spot_index = column_names.index("USD Spot") + 1
            usd_total_index = column_names.index("USD Total") + 1
            reason_index = column_names.index("Reason") + 1

            row = 4
            for conversion in self.conversions:
                
                if conversion.symbol != asset:
                    continue

                conversions_sheet.cell(row=row, column=in_trans_type_index, value=conversion.input_trans_type)
                conversions_sheet.cell(row=row, column=out_trans_type_index, value=conversion.output_trans_type)
                conversions_sheet.cell(row=row, column=symbol_index, value=conversion.symbol)
                conversions_sheet.cell(row=row, column=time_stamp_index, value=conversion.time_stamp)
                conversions_sheet.cell(row=row, column=quantity_index, value=conversion.quantity)
                
                conversions_sheet.cell(row=row, column=usd_spot_index, value=conversion.usd_spot)
                conversions_sheet.cell(row=row, column=usd_spot_index).number_format = '"$"#,##0.00_-'
                
                conversions_sheet.cell(row=row, column=usd_total_index, value=conversion.usd_total)
                conversions_sheet.cell(row=row, column=usd_total_index).number_format = '"$"#,##0.00_-'
                
                conversions_sheet.cell(row=row, column=reason_index, value=conversion.reason)

                row += 1

            if row == 4:
                workbook.remove(conversions_sheet)


            # Gainz Sheet
            column_names = []
            for cell in l_sheet[1]:
                column_names.append(cell.value)

            sell_date_index = column_names.index("Sell Date") + 1
            sell_id_index = column_names.index("Sell ID") + 1
            sell_quantity_index = column_names.index("Sell Quantity") + 1
            sell_unlinked_index = column_names.index("Sell Unlinked") + 1
            sell_usd_total_index = column_names.index("Sell USD Total") + 1
            sell_usd_spot_index = column_names.index("Sell Spot USD") + 1
            sell_multi_link_index = column_names.index("Sell Multi-Link") + 1

            buy_link_usd_index = column_names.index("Link Buy in USD") + 1
            link_id_index = column_names.index("Link ID") + 1
            link_symbol_index = column_names.index("Link Asset") + 1
            link_quantity_index = column_names.index("Link Quantity") + 1
            link_profit_loss_index = column_names.index("Link Profit Loss") + 1
            sell_link_usd_index = column_names.index("Link Sell in USD") + 1
            date_acquired_index = column_names.index("Date Acquired") + 1

            buy_multi_link_index = column_names.index("Buy Multi-Link") + 1
            buy_date_index = column_names.index("Buy Date") + 1
            buy_id_index = column_names.index("Buy ID") + 1
            buy_quantity_index = column_names.index("Buy Quantity") + 1
            buy_unlinked_index = column_names.index("Buy Unlinked") + 1
            buy_usd_total_index = column_names.index("Buy USD Total") + 1
            buy_usd_spot_index = column_names.index("Buy Spot USD") + 1

            years = set()
            for link in self.links:
                if link.symbol != asset:
                    continue

                years.add(link.sell.time_stamp.year)

            for year in years:
                
                sheetname = f'{year} {asset} Gains'
                links_sheet = workbook.copy_worksheet(l_sheet)
                links_sheet.title = sheetname
                        
                row = 2
                profit_loss_total = 0.0
                for link in self.links:
                    
                    if link.symbol != asset:
                        continue
                        
                    if link.sell.time_stamp.year != year:
                        continue

                    if link.quantity <= 0.00000001:
                        continue

                    profit_loss_total += float(link.profit_loss)
                    
                    links_sheet.cell(row=row, column=link_symbol_index, value=link.sell.symbol)
                    links_sheet.cell(row=row, column=link_id_index, value=link.id)
                    links_sheet.cell(row=row, column=buy_id_index, value=link.buy.id)
                    links_sheet.cell(row=row, column=sell_id_index, value=link.sell.id)
                    links_sheet.cell(row=row, column=buy_date_index, value=link.buy.time_stamp)
                    links_sheet.cell(row=row, column=buy_quantity_index, value=link.buy.quantity)
                    links_sheet.cell(row=row, column=buy_unlinked_index, value=link.buy.unlinked_quantity)
                    
                    links_sheet.cell(row=row, column=buy_usd_spot_index, value=link.buy.usd_spot)
                    links_sheet.cell(row=row, column=buy_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=buy_usd_total_index, value=link.buy.usd_total)
                    links_sheet.cell(row=row, column=buy_usd_total_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=buy_link_usd_index, value=link.link_buy_price)
                    links_sheet.cell(row=row, column=buy_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=link_quantity_index, value=link.quantity)
                    
                    links_sheet.cell(row=row, column=link_profit_loss_index, value=link.profit_loss)
                    links_sheet.cell(row=row, column=link_profit_loss_index).number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

                    links_sheet.cell(row=row, column=sell_link_usd_index, value=link.link_sell_price)
                    links_sheet.cell(row=row, column=sell_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_date_index, value=link.sell.time_stamp)
                    links_sheet.cell(row=row, column=sell_quantity_index, value=link.sell.quantity)
                    links_sheet.cell(row=row, column=sell_unlinked_index, value=link.sell.unlinked_quantity)

                    links_sheet.cell(row=row, column=sell_usd_spot_index, value=link.sell.usd_spot)
                    links_sheet.cell(row=row, column=sell_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_usd_total_index, value=link.sell.usd_total)
                    links_sheet.cell(row=row, column=sell_usd_total_index).number_format = '"$"#,##0.00_-'
                    
                    links_sheet.cell(row=row, column=sell_multi_link_index, value=link.sell.multi_link)
                    links_sheet.cell(row=row, column=buy_multi_link_index, value=link.buy.multi_link)
                    
                    row += 1

                for trans in self.transactions:
                    
                    if trans.symbol != asset:
                        continue

                    if trans.trans_type != 'sell':
                        continue

                    if trans.time_stamp.year != year:
                        continue

                    if trans.unlinked_quantity <= 0.00000001:
                        continue

                    profit_loss_total += float(trans.unlinked_quantity * trans.usd_spot)

                    links_sheet.cell(row=row, column=link_symbol_index, value="N/A")
                    links_sheet.cell(row=row, column=link_id_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_id_index, value="N/A")
                    links_sheet.cell(row=row, column=sell_id_index, value=trans.id)
                    links_sheet.cell(row=row, column=buy_date_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_quantity_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_unlinked_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_usd_spot_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_usd_total_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_link_usd_index, value="N/A")
                    links_sheet.cell(row=row, column=link_quantity_index, value=trans.unlinked_quantity)

                    links_sheet.cell(row=row, column=link_profit_loss_index, value=(trans.unlinked_quantity * trans.usd_spot))
                    links_sheet.cell(row=row, column=link_profit_loss_index).number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                    
                    links_sheet.cell(row=row, column=sell_link_usd_index, value=trans.usd_total)
                    links_sheet.cell(row=row, column=sell_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_date_index, value=trans.time_stamp)
                    links_sheet.cell(row=row, column=sell_quantity_index, value=trans.quantity)
                    links_sheet.cell(row=row, column=sell_unlinked_index, value=trans.unlinked_quantity)
                    links_sheet.cell(row=row, column=sell_usd_spot_index, value=trans.usd_spot)
                    links_sheet.cell(row=row, column=sell_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_usd_total_index, value=trans.usd_total)
                    links_sheet.cell(row=row, column=sell_usd_total_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_multi_link_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_multi_link_index, value="N/A")

                    row += 1
                    

                row += 2

                links_sheet.cell(row=row, column=link_profit_loss_index, value="Profit/Loss Total: ${:,.2f}".format(profit_loss_total)) 

                if row == 4:
                    workbook.remove(links_sheet)

            # All Transactions sheet
            sheetname = f'{asset} Transactions'
            all_trans_sheet = workbook.copy_worksheet(a_sheet)
            all_trans_sheet.title = sheetname
            
            column_names = []
            for cell in all_trans_sheet[1]:
                column_names.append(cell.value)

            id_index = column_names.index("Transaction ID") + 1
            symbol_index = column_names.index("Symbol") + 1
            trans_type_index = column_names.index("Transaction Type") + 1
            time_stamp_index = column_names.index("Time Stamp") + 1 
            quantity_index = column_names.index("Quantity") + 1
            links_index = column_names.index("Links") + 1
            unlinked_index = column_names.index("Unlinked") + 1
            usd_spot_index = column_names.index("USD Spot") + 1
            usd_total_index = column_names.index("USD Total") + 1
            source_index = column_names.index("Source") + 1
            
            row = 2
            for trans in self.transactions:

                if trans.symbol != asset:
                    continue

                all_trans_sheet.cell(row=row, column=id_index, value=trans.id)
                all_trans_sheet.cell(row=row, column=symbol_index, value=trans.symbol)
                all_trans_sheet.cell(row=row, column=trans_type_index, value=trans.trans_type)
                all_trans_sheet.cell(row=row, column=time_stamp_index, value=trans.time_stamp)
                all_trans_sheet.cell(row=row, column=quantity_index, value=trans.quantity)
                all_trans_sheet.cell(row=row, column=unlinked_index, value=trans.unlinked_quantity)

                all_trans_sheet.cell(row=row, column=usd_spot_index, value=trans.usd_spot)
                all_trans_sheet.cell(row=row, column=usd_spot_index).number_format = '"$"#,##0.00_-'

                all_trans_sheet.cell(row=row, column=usd_total_index, value=trans.usd_total)
                all_trans_sheet.cell(row=row, column=usd_total_index).number_format = '"$"#,##0.00_-'

                all_trans_sheet.cell(row=row, column=source_index, value=trans.source)

                if len(trans.links) > 0:
                    all_trans_sheet.cell(row=row, column=links_index, value=str(trans.links))

                row += 1

            if row == 2:
                workbook.remove(all_trans_sheet)


            # Links Sheet
            sheetname = f'{asset} Stats'
            asset_stats_sheet = workbook.copy_worksheet(s_sheet)
            asset_stats_sheet.title = sheetname


            date_range = {
                'start_date': '',
                'end_date': ''
            }

            date_range = get_transactions_date_range(self, date_range)

            # get stats table data 
            stats_table_data = get_stats_table_data_range(self, date_range)

            # get stats for selected asset
            asset_stats = None
            for a in stats_table_data:
                if a['symbol'] == asset:
                    asset_stats = a
                    break
                
           
            # print(asset_stats)

            # Create detailed stats table data
            detailed_stats = [
                ["Quantity Purchased", asset_stats['total_purchased_quantity']],
                ["Number of Buys", asset_stats["num_buys"]],
                ["Number of Sells", asset_stats["num_sells"]],
                ["Number of Links", asset_stats["num_links"]],
                ["Average Buy Price", asset_stats["average_buy_price"]],
                ["Average Sell Price", asset_stats["average_sell_price"]],
                ["Quantity Sold", asset_stats['total_sold_quantity']],
                ["Quantity Sold Unlinked", asset_stats['total_sold_unlinked_quantity']],
                ["Quantity Purchased Unlinked, AKA The HODL", asset_stats['total_purchased_unlinked_quantity']],
                ["Quantity Purchased in USD", asset_stats['total_purchased_usd']],
                ["Quantity Sold in USD", asset_stats['total_sold_usd']],
                ["Profit / Loss in USD* (Valid when Quantity Sold Unlinked is 0)", asset_stats['total_profit_loss']],
            ]

            row = 2
            for i in detailed_stats:
                asset_stats_sheet.cell(row=row, column=1, value=i[0])
                asset_stats_sheet.cell(row=row, column=1).number_format = '"$"#,##0.00_-'
                asset_stats_sheet.cell(row=row, column=2, value=i[1])
                asset_stats_sheet.cell(row=row, column=2).number_format = '"$"#,##0.00_-'
                
                row += 1

        
        workbook.remove(a_sheet)
        workbook.remove(c_sheet)
        workbook.remove(l_sheet)
        workbook.remove(s_sheet)
        workbook.remove(t8949_sheet)
            
            
        workbook.save(save_as_filename)
        print(f"Saving to {save_as_filename}")

        return save_as_filename


        
    def import_transaction(self, type, timestamp, quantity):

        if type.lower() == 'buy':
            trans = Buy(trans_type=type, time_stamp=timestamp, quantity=quantity, symbol='BTC', source='Gainz App')
        elif type.lower() == 'sell':
            trans = Sell(trans_type=type, time_stamp=timestamp, quantity=quantity, symbol='BTC', source='Gainz App')
        
        self.transactions.append(trans)
        print(len(self.transactions))

        return trans


    def import_transactions(self, contents, filename):

        if type(contents) is str:
            content_type, content_string = contents.split(',')
            decoded = base64.b64decode(content_string)
            try:
                if 'csv' in filename:
                    if 'cash_app' in filename.lower():
                        df = pd.read_csv(contents, on_bad_lines='skip')
                    elif 'coinbase' in filename.lower():
                        df = pd.read_csv(contents, on_bad_lines='skip', skip_blank_lines=False, header=7)
                elif 'xls' in filename or 'xlsx' in filename:
                    # Assume that the user uploaded an excel file
                    df = pd.read_excel(io.BytesIO(decoded))
                
            except Exception as e:
                print(e)
        else:
            try:
                if 'csv' in filename:
                    if 'cash_app' in filename.lower():
                        df = pd.read_csv(contents, on_bad_lines='skip')
                    elif 'coinbase' in filename.lower():
                        df = pd.read_csv(contents, on_bad_lines='skip', skip_blank_lines=False, header=7)
                elif 'xls' in filename or 'xlsx' in filename:
                    # Assume that the user uploaded an excel file
                    df = pd.read_excel(contents, error_bad_lines=False)
            
            except Exception as e:
                print(e)

        # print(f"Base Dir in import transactions {basedir}")
        # pathlib.Path(os.path.join(basedir, "saves")).mkdir(parents=True, exist_ok=True) 

        save_as_filename = os.path.join("saves", f"saved_{strftime('Y%Y-M%m-D%d_H%H-M%M-S%S')}.xlsx")
        

        # trans_df = pd.read_csv(filename, header=3)
        trans_df = df

        convert_df = None

        # Import from cashapp csv
        if 'Date' in trans_df.columns:

            print("Date Found in columns")
            trans_df.rename(columns={'Date': 'Timestamp', 'Asset Type': 'Asset', 'Asset Amount': 'Quantity Transacted', 'Asset Price': 'USD Spot Price at Transaction'}, inplace=True)

            trans_df['Timestamp'] = pd.to_datetime(trans_df['Timestamp'], infer_datetime_format=True, utc=True) 
            trans_df['Timestamp'] = trans_df['Timestamp'].dt.tz_localize(None)

            trans_df['USD Spot Price at Transaction'] = trans_df['USD Spot Price at Transaction'].replace('[\$,]', '', regex=True).astype(float)
            trans_df['Source'] = filename

            sell_df = trans_df[trans_df['Transaction Type'] == 'Bitcoin Sale'].copy()
            buy_df = trans_df[trans_df['Transaction Type'] == 'Bitcoin Buy'].copy()
            send_df = trans_df[trans_df['Transaction Type'] == 'Bitcoin Withdrawal'].copy()
            receive_df = trans_df[trans_df['Transaction Type'] == 'Bitcoin Deposit'].copy()
            convert_df = None
            

        #import from coinbase csv
        else:
            
            # print(trans_df.columns)
            trans_df['Timestamp'] = pd.to_datetime(trans_df['Timestamp']) 
            trans_df['Timestamp'] = trans_df['Timestamp'].dt.tz_localize(None)
            trans_df['Source'] = filename

            sell_df = trans_df[trans_df['Transaction Type'] == 'Sell'].copy()
            buy_df = trans_df[trans_df['Transaction Type'] == 'Buy'].copy()
            send_df = trans_df[trans_df['Transaction Type'] == 'Send'].copy()
            receive_df = trans_df[trans_df['Transaction Type'] == 'Receive'].copy()
            convert_df = trans_df[trans_df['Transaction Type'] == 'Convert'].copy()

        sell_df.reset_index(inplace=True)
        buy_df.reset_index(inplace=True)
        send_df.reset_index(inplace=True)
        receive_df.reset_index(inplace=True)


        sell_df.sort_values(by='Timestamp', inplace=True)
        buy_df.sort_values(by='Timestamp', inplace=True)
        send_df.sort_values(by='Timestamp', inplace=True)
        receive_df.sort_values(by='Timestamp', inplace=True)

        # Objects >
        sells = []
        buys = []
        sends = []
        receives = []
        assets = []


        ## Create Objects
        # Sells
        for index, row in sell_df.iterrows():
            sells.append(Sell(symbol=row['Asset'], quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction'], source=row['Source']))

        # Buys
        for index, row in buy_df.iterrows():
            buys.append(Buy(symbol=row['Asset'], quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction'], source=row['Source']))

        # Sends
        for index, row in send_df.iterrows():
            sends.append(Send(symbol=row['Asset'], quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction'], source=row['Source']))

        # Receives
        for index, row in receive_df.iterrows():
            receives.append(Receive(symbol=row['Asset'], quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction'], source=row['Source']))

        
        if convert_df is not None:
            convert_df.reset_index(inplace=True)
            for index, row in convert_df.iterrows():

                symbol=row['Asset']
                quantity=row['Quantity Transacted']
                time_stamp=row['Timestamp']
                usd_spot=row['USD Spot Price at Transaction']
                source=row['Source']

                sells.append(Sell(symbol=row['Asset'], quantity=row['Quantity Transacted'], time_stamp=row['Timestamp'], usd_spot=row['USD Spot Price at Transaction'], source=f"{row['Source']} - Convert Sell"))

                usd_total = row['USD Total (inclusive of fees)']

        
                notes = row['Notes']

                buy_quantity = float(notes.split()[4])
                buy_asset = notes.split()[5]

                buy_usd_spot = usd_total / buy_quantity

                buys.append(Buy(symbol=buy_asset, quantity=buy_quantity, time_stamp=row['Timestamp'], usd_spot=buy_usd_spot, source=f"{row['Source']} - Convert Buy" ))

        if self.transactions:
            
            deduped_transactions = set()

            imported_transactions =  buys + sells + sends + receives

            for trans in self.transactions:
                deduped_transactions.add(trans)

            for trans in imported_transactions:
                deduped_transactions.add(trans)
                
            self.transactions = list(deduped_transactions)
        
        else:
            
            self.transactions = buys + sells + sends + receives

        
        for asset in self.assets:
            asset_obj = Asset(symbol=asset)
            assets.append(asset_obj)
        
        self.asset_objects = assets


        # self.save(description="Imported Transactions from CSV")
        # Convert Objects to DataFrame
        trans_df = pd.DataFrame([vars(s) for s in self.transactions])
        conversion_df = pd.DataFrame([vars(s) for s in self.conversions])
        asset_df = pd.DataFrame([vars(s) for s in self.asset_objects])
        links_df = pd.DataFrame([vars(s) for s in self.links])

        # Write Objects to excel file
        
        with pd.ExcelWriter(save_as_filename) as writer:
            trans_df.to_excel(writer, sheet_name="All Transactions")
            conversion_df.to_excel(writer, sheet_name="Conversions")
            asset_df.to_excel(writer, sheet_name="Assets")
            links_df.to_excel(writer, sheet_name="Links")


        print(f"\nImported Transactions Saving to {save_as_filename}\n")

        return save_as_filename

    def first_transaction_date(self, asset=None):
        all_trans = {}
        
        for trans in self.transactions:

            # If asset is provided skip others
            if asset is not None:
                if trans.symbol != asset:
                    continue
            
            # Create key val for symbol
            if trans.symbol not in all_trans.keys():
                all_trans[trans.symbol] = []

            all_trans[trans.symbol].append(trans)

        # Sort By Time Stamp
        for key in all_trans.keys():
            all_trans[key].sort(key=lambda x: x.time_stamp)

        # Extract first transaction Date
        first_time_stamps = {}
        for key in all_trans.keys():
            first_time_stamps[key] = all_trans[key][0].time_stamp
            
            # print(first_time_stamps)
            # print(f"First Transaction Date for {key}: {all_trans[key][0].time_stamp}")

        return first_time_stamps

    def last_transaction_date(self, asset=None):
        all_trans = {}
        
        # Sort into Buys a Sells
        for trans in self.transactions:

            # If asset is provided only auto-link symbol provided
            if asset is not None:
                if trans.symbol != asset:
                    continue
            
            # Create key val for symbol
            if trans.symbol not in all_trans.keys():
                all_trans[trans.symbol] = []


            all_trans[trans.symbol].append(trans)

        
        # Sort By Time Stamp
        for key in all_trans.keys():
            all_trans[key].sort(key=lambda x: x.time_stamp)

        # Extract Last transaction Date
        last_time_stamps = {}
        for key in all_trans.keys():
            last_time_stamps[key] = all_trans[key][-1].time_stamp
            
        # print(last_time_stamps)
        # print(f"Last Transaction Date for {key}: {all_trans[key][-1].time_stamp}")

        return last_time_stamps


if __name__ == "__main__":

    transactions = Transactions()

  
